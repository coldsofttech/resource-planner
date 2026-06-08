from __future__ import annotations

import io
import unittest

import openpyxl
from exportcore.xlsx_export import export_to_xlsx


class ExportToXlsxEmptyRowsTest(unittest.TestCase):
    def setUp(self):
        self.result = export_to_xlsx([])

    def test_returns_bytes(self):
        self.assertIsInstance(self.result, bytes)

    def test_output_is_valid_xlsx(self):
        wb = openpyxl.load_workbook(io.BytesIO(self.result))
        self.assertIsNotNone(wb)

    def test_sheet_title_defaults_to_export(self):
        wb = openpyxl.load_workbook(io.BytesIO(self.result))
        self.assertEqual(wb.active.title, "Export")

    def test_no_header_row_for_empty_rows(self):
        wb = openpyxl.load_workbook(io.BytesIO(self.result))
        ws = wb.active
        self.assertIsNone(ws.cell(row=1, column=1).value)


class ExportToXlsxSingleRowTest(unittest.TestCase):
    def setUp(self):
        self.rows = [{"Name": "Alice", "Role": "Engineer"}]
        self.result = export_to_xlsx(self.rows)
        self.ws = openpyxl.load_workbook(io.BytesIO(self.result)).active

    def test_returns_bytes(self):
        self.assertIsInstance(self.result, bytes)

    def test_header_row_has_correct_values(self):
        self.assertEqual(self.ws.cell(row=1, column=1).value, "Name")
        self.assertEqual(self.ws.cell(row=1, column=2).value, "Role")

    def test_data_row_has_correct_values(self):
        self.assertEqual(self.ws.cell(row=2, column=1).value, "Alice")
        self.assertEqual(self.ws.cell(row=2, column=2).value, "Engineer")

    def test_total_rows_is_header_plus_data(self):
        self.assertEqual(self.ws.max_row, 2)

    def test_total_columns_matches_field_count(self):
        self.assertEqual(self.ws.max_column, 2)


class ExportToXlsxMultipleRowsTest(unittest.TestCase):
    def setUp(self):
        self.rows = [
            {"Name": "Alice", "Role": "Engineer"},
            {"Name": "Bob", "Role": "Designer"},
            {"Name": "Carol", "Role": "Manager"},
        ]
        self.result = export_to_xlsx(self.rows)
        self.ws = openpyxl.load_workbook(io.BytesIO(self.result)).active

    def test_all_data_rows_written(self):
        self.assertEqual(self.ws.max_row, 4)

    def test_data_values_are_correct(self):
        names = [self.ws.cell(row=r, column=1).value for r in range(2, 5)]
        self.assertIn("Alice", names)
        self.assertIn("Bob", names)
        self.assertIn("Carol", names)

    def test_column_order_matches_first_row_keys(self):
        self.assertEqual(self.ws.cell(row=1, column=1).value, "Name")
        self.assertEqual(self.ws.cell(row=1, column=2).value, "Role")


class ExportToXlsxHeaderStylingTest(unittest.TestCase):
    def setUp(self):
        self.rows = [{"Title": "Project Alpha", "Status": "Active"}]
        self.result = export_to_xlsx(self.rows)
        self.ws = openpyxl.load_workbook(io.BytesIO(self.result)).active

    def test_header_font_is_bold(self):
        self.assertTrue(self.ws.cell(row=1, column=1).font.bold)

    def test_header_font_color_is_white(self):
        font_color = self.ws.cell(row=1, column=1).font.color.rgb
        self.assertEqual(font_color.upper(), "FFFFFFFF")

    def test_header_fill_color_matches_accent(self):
        fill_color = self.ws.cell(row=1, column=1).fill.fgColor.rgb
        self.assertEqual(fill_color.upper(), "FF4B65D9")

    def test_header_row_height_is_set(self):
        self.assertEqual(self.ws.row_dimensions[1].height, 22)


class ExportToXlsxAlternatingRowFillTest(unittest.TestCase):
    def setUp(self):
        self.rows = [
            {"Col": "A"},
            {"Col": "B"},
            {"Col": "C"},
        ]
        self.result = export_to_xlsx(self.rows)
        self.ws = openpyxl.load_workbook(io.BytesIO(self.result)).active

    def test_even_data_rows_have_light_fill(self):
        fill = self.ws.cell(row=2, column=1).fill.fgColor.rgb
        self.assertEqual(fill.upper(), "FFEEF2FF")

    def test_odd_data_rows_have_no_fill(self):
        fill = self.ws.cell(row=3, column=1).fill.fgColor.rgb
        self.assertNotEqual(fill.upper(), "FFEEF2FF")


class ExportToXlsxColumnWidthTest(unittest.TestCase):
    def test_short_header_uses_minimum_width_of_14(self):
        rows = [{"AB": "x"}]
        result = export_to_xlsx(rows)
        ws = openpyxl.load_workbook(io.BytesIO(result)).active
        self.assertEqual(ws.column_dimensions["A"].width, 14)

    def test_long_header_uses_header_length_plus_four(self):
        long_header = "VeryLongColumnHeaderName"
        rows = [{long_header: "val"}]
        result = export_to_xlsx(rows)
        ws = openpyxl.load_workbook(io.BytesIO(result)).active
        self.assertEqual(ws.column_dimensions["A"].width, len(long_header) + 4)


class ExportToXlsxMissingKeyHandlingTest(unittest.TestCase):
    def test_missing_key_in_subsequent_row_produces_empty_string(self):
        rows = [
            {"Name": "Alice", "Role": "Engineer"},
            {"Name": "Bob"},
        ]
        result = export_to_xlsx(rows)
        ws = openpyxl.load_workbook(io.BytesIO(result)).active
        self.assertIsNone(ws.cell(row=3, column=2).value)


class ExportToXlsxSheetTitleTest(unittest.TestCase):
    def test_custom_title_used_as_sheet_name(self):
        result = export_to_xlsx([{"Name": "Alice"}], title="Teams")
        ws = openpyxl.load_workbook(io.BytesIO(result)).active
        self.assertEqual(ws.title, "Teams")

    def test_default_title_is_export(self):
        result = export_to_xlsx([{"Name": "Alice"}])
        ws = openpyxl.load_workbook(io.BytesIO(result)).active
        self.assertEqual(ws.title, "Export")


class ExportToXlsxLargeDatasetTest(unittest.TestCase):
    def test_handles_100_rows_without_error(self):
        rows = [{"ID": str(i), "Value": f"item_{i}"} for i in range(100)]
        result = export_to_xlsx(rows)
        ws = openpyxl.load_workbook(io.BytesIO(result)).active
        self.assertEqual(ws.max_row, 101)

    def test_handles_many_columns_without_error(self):
        headers = {f"Col_{i}": f"val_{i}" for i in range(20)}
        result = export_to_xlsx([headers])
        ws = openpyxl.load_workbook(io.BytesIO(result)).active
        self.assertEqual(ws.max_column, 20)
