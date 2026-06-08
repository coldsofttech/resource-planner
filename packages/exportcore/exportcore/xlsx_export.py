from __future__ import annotations

import io


def export_to_xlsx(rows: list[dict[str, str]], *, title: str = "Export") -> bytes:
    """Render rows to XLSX bytes with a branded header and alternating rows.

    Args:
        rows:  Pre-formatted list of dicts where keys are display column names.
        title: Worksheet tab name (defaults to "Export").

    Returns:
        Raw XLSX bytes.
    """
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    # oklch(0.55 0.18 270) → #4B65D9
    _PRIMARY = "FF4B65D9"
    _LIGHT_ROW = "FFEEF2FF"
    _BORDER = "FFE0E7FF"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title

    headers = list(rows[0].keys()) if rows else []

    header_fill = PatternFill(
        start_color=_PRIMARY, end_color=_PRIMARY, fill_type="solid"
    )
    header_font = Font(color="FFFFFFFF", bold=True, size=10)
    alt_fill = PatternFill(
        start_color=_LIGHT_ROW, end_color=_LIGHT_ROW, fill_type="solid"
    )
    thin = Side(style="thin", color=_BORDER)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row.get(header, ""))
            cell.font = Font(size=9)
            cell.alignment = Alignment(vertical="center")
            cell.border = border
            if row_idx % 2 == 0:
                cell.fill = alt_fill

    for col_idx, header in enumerate(headers, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max(
            len(header) + 4, 14
        )

    ws.row_dimensions[1].height = 22

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
